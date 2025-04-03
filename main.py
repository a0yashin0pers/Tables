import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def parse_text_file(file_path):
    """
    Считываем текстовый файл и формируем структуру tables:
    [
      {
        "name": "Название_таблицы",
        "header": [...],
        "rows": [
           [...],
           [...],
           ...
        ]
      },
      ...
    ]
    По условию заменяем все точки на запятые.
    """
    with open(file_path, encoding='utf-8') as f:
        lines = f.read().splitlines()

    tables = []
    current_table = None

    for line in lines:
        # Заменяем точки на запятые
        line = line.replace('.', ',').strip()
        if not line:
            # Пустая строка — значит закончилась предыдущая таблица
            current_table = None
            continue

        if '#' not in line:
            # Это строка с названием новой таблицы
            current_table = {"name": line, "header": None, "rows": []}
            tables.append(current_table)
        else:
            # Это строка вида: "заголовок1&заголовок2&...#значение1&значение2&..."
            header_part, data_part = [p.strip() for p in line.split('#', 1)]
            cols = [x.strip() for x in header_part.split('&')]
            vals = [x.strip() for x in data_part.split('&')]

            # Если не было явно заданного названия таблицы
            if current_table is None:
                current_table = {"name": "Без названия", "header": cols, "rows": []}
                tables.append(current_table)

            # Если у текущей таблицы ещё нет заголовка, назначаем его
            if current_table["header"] is None:
                current_table["header"] = cols

            # Добавляем строку значений
            current_table["rows"].append(vals)

    return tables

def write_tables_to_excel(tables, output_file):
    """
    Записываем данные всех таблиц в один лист Excel ("Tables") подряд.
      1) Для каждой таблицы пишем строку с названием, строку с заголовками, строки с данными, а затем пустую строку.
      2) Объединяем ячейки в столбце, если подряд идущие одинаковые данные (и не пустые).
      3) Объединяем ячейку с названием таблицы по всей ширине (количеству столбцов таблицы).
      4) Регулируем ширину столбца по длине заголовка.
      5) Переводим «числовые» строки в настоящие числа (float или int).
    """
    # Сформируем общий список строк (sheet_data) для одного листа
    sheet_data = []
    for table in tables:
        # 1) строка с названием таблицы
        sheet_data.append([table["name"]])
        # 2) строка с заголовками (если есть)
        if table["header"]:
            sheet_data.append(table["header"])
        else:
            sheet_data.append([])
        # 3) строки с данными
        sheet_data.extend(table["rows"])
        # 4) пустая строка-разделитель
        sheet_data.append([])

    # Выравниваем по максимальному числу столбцов
    max_cols = max((len(row) for row in sheet_data), default=0)
    normalized = [row + [""] * (max_cols - len(row)) for row in sheet_data]

    df = pd.DataFrame(normalized)

    # Пишем в Excel через pd.ExcelWriter
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Tables", header=False, index=False)

        ws = writer.book["Tables"]

        # 1) Пробуем конвертировать значения из строк в числа
        convert_strings_to_numbers(ws)

        # 2) Объединяем ячейки по столбцам, если подряд идущие одинаковые (и не пустые)
        merge_cells_in_worksheet(ws)

        # 3) Объединяем ячейку с названием таблицы по всей ширине
        merge_table_names(ws, tables)

        # 4) Регулируем ширину столбцов по длине заголовков
        set_column_widths_by_header(ws, tables)

        # Закрытие writer – при выходе из with всё автоматически сохранится

    print(f"Excel файл сохранён как '{output_file}'")

def convert_strings_to_numbers(ws):
    """
    Проходим по всем ячейкам листа. Если значение (cell.value) – непустая строка,
    пытаемся считать его как число (учитывая, что дробная часть отделена запятой).
    Если получается – записываем в ячейку это число, ставим формат без апострофа.
    Если нет – оставляем как есть (строку).
    """
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str):
                val_str = val.strip()
                if val_str:
                    # Пробуем заменить запятую на точку и прочитать как float
                    maybe_num = val_str.replace(',', '.')
                    try:
                        number = float(maybe_num)
                        # Если после запятой нет десятичной части, пусть будет int
                        # (например, "10" или "10,0")
                        if number.is_integer():
                            cell.value = int(number)
                            cell.number_format = "0"
                        else:
                            cell.value = number
                            cell.number_format = "0.00"

                        cell.data_type = 'n'  # Принудительно указываем, что это число
                    except ValueError:
                        # Не удалось преобразовать в число – оставляем строку
                        pass

def merge_cells_in_worksheet(ws):
    """
    Объединяем подряд идущие ячейки в одном столбце, если их значения совпадают (и не пустые).
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        row_start = 1
        current_value = ws.cell(row=row_start, column=col).value

        for row in range(row_start + 1, max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            # Проверяем совпадение и "не пустоту"
            if cell_value != current_value or current_value in (None, ""):
                # Если цепочка одинаковых непустых значений прервалась - мёржим предыдущий диапазон
                if current_value not in (None, "") and (row - 1 > row_start):
                    ws.merge_cells(start_row=row_start, start_column=col,
                                   end_row=row - 1, end_column=col)
                # Обновляем начало нового диапазона
                row_start = row
                current_value = cell_value

        # В конце столбца проверяем, не нужно ли домёржить
        if current_value not in (None, "") and (row_start < max_row):
            ws.merge_cells(start_row=row_start, start_column=col,
                           end_row=max_row, end_column=col)

def merge_table_names(ws, tables):
    """
    После записи всех таблиц подряд (с одной пустой строкой разделителя), 
    объединяем ячейку с названием таблицы на всю «ширину» текущей таблицы.
    
    Ширину для конкретной таблицы определяем:
      - Если есть header, то это len(header),
      - Иначе берём макс. длину строки в rows,
      - Если и того нет, пусть будет 1.
    """
    row_counter = 1  # начнём с 1, т.к. в openpyxl строки нумеруются с 1
    for table in tables:
        # Ширина таблицы
        header_width = len(table["header"]) if table["header"] else 0
        rows_width = max((len(r) for r in table["rows"]), default=0)
        table_width = max(header_width, rows_width, 1)

        # Название таблицы лежит в строке row_counter (первая строка для данной таблицы)
        ws.merge_cells(start_row=row_counter, start_column=1,
                       end_row=row_counter, end_column=table_width)

        # Подсчитаем, сколько строк занимает таблица:
        #  - 1 строка: название
        #  - 1 строка: заголовок (или пустая, если заголовка нет)
        #  - N строк: rows
        #  - 1 пустая строка
        total_rows_for_table = 1 + 1 + len(table["rows"]) + 1
        row_counter += total_rows_for_table

def set_column_widths_by_header(ws, tables, default_width=10):
    """
    Устанавливаем ширину столбца по длине названия столбца (header),
    только если длина названия больше, чем default_width.
    """
    row_offset = 0
    for table in tables:
        # Строка с названием таблицы = row_offset + 1
        # Строка с заголовками = row_offset + 2
        header_row_idx = row_offset + 2

        # Если заголовок есть, меняем ширину столбцов
        if table["header"]:
            for col_idx, col_name in enumerate(table["header"], start=1):
                if col_name is None:
                    continue
                length = len(str(col_name))
                if length > default_width:
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = length

        # Сколько строк занимает таблица: 1 (название) + 1 (заголовок) + len(rows) + 1 (пустая)
        total_rows_for_table = 1 + 1 + len(table["rows"]) + 1
        row_offset += total_rows_for_table

if __name__ == "__main__":
    input_file = "results.txt"
    output_file = "output.xlsx"
    tables = parse_text_file(input_file)
    write_tables_to_excel(tables, output_file)